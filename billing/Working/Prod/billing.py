from flask import Flask, render_template, request, jsonify
from flask import Response
import mysql.connector
import json
import requests
import random
import os
import openpyxl
import calendar
from datetime import datetime
from werkzeug.wrappers import response
import subprocess
import sys

app = Flask(__name__)


@app.route('/')
def index():
    return render_template('index.html')
    # Add Navigiation bar to our APIs


@app.route('/bills')
def bills():
    return render_template('bills.html')
    # Add Navigiation bar to our APIs

@app.route('/bills/<provider_id>')
def bills_spes(provider_id):
    return render_template('bills_spec.html')
    # Add Navigiation bar to our APIs

@app.route('/health')
def health():
    return render_template('health.html')
@app.route('/api/health')
def healtho():
    try:
        mycursor.execute("use billdb")
        return Response({"Connection to the database is healthy"}, status=200)
    except:
        return Response({"Connection to the darabase is not healthy"},status=500)

@app.route('/providers')
def provider():
    return render_template('providers.html')


@app.route('/rates')
def rates():
    mydir = os.listdir("/in")
    return render_template('rates.html', mydir=mydir)


@app.route('/trucks')
def truck():
    return render_template('trucks.html')


@app.route('/trucks/<truck_number>')
def truck_id(truck_number):
    return render_template('truck_id.html')


@app.route('/api/providers', methods=['GET', 'POST'])
def providers():
    if request.method == 'POST':
        provider = request.form['provider']
        mycursor = billingdb.cursor()
        mycursor.execute("USE billdb")
        mycursor.execute(f"SELECT name from Provider where name='{str(provider)}'")
        results = mycursor.fetchall()
        if not results:
            switch = True
            while switch:
                prov_id = random.randint(1, 999999)
                jsonprovider = {'id': str(prov_id), 'name': str(provider)}
                try:
                    mycursor.execute("USE billdb")
                    mycursor.execute(f"INSERT INTO Provider(id, name) VALUES('{str(prov_id)}', '{str(provider)}')")
                    switch = False
                except mysql.connector.errors.IntegrityError:
                    continue
                except:
                    return Response('error', status=400)
                return jsonify(jsonprovider)
        else:
            return Response('name exist', status=400)
    if request.method == 'GET':
        return Response("enter provider name:", mimetype='text/plain')


@app.route('/api/rates', methods=['POST', 'GET'])
def ratespost():
    if request.method == 'GET':
        mycursor = billingdb.cursor()
        mycursor.execute("""select * from Rates""")
        row_headers = [x[0] for x in mycursor.description]
        rv = mycursor.fetchall()
        json_data = []
        for result in rv:
            json_data.append(dict(zip(row_headers, result)))
        jsonout = json.dumps(json_data)
        return jsonout  # Check with chris if the JSON type is required on response
    if request.method == 'POST':
        mycursor = billingdb.cursor()
        filename = request.form['msgfile']
        try:
            wrkbk = openpyxl.load_workbook(f"/in/{filename}")
        except openpyxl.utils.exceptions.InvalidFileException:
            return "no such a file"

        sh = wrkbk.active
        semilist = []
        mylist = []
        for i in range(2, sh.max_row + 1):
            mylist.append(semilist)
            semilist = []
            for j in range(1, sh.max_column + 1):
                cell_obj = sh.cell(row=i, column=j)
                semilist.append(str(cell_obj.value))
        mylist.pop(0)
        for i in mylist:
            try:
                mycursor.execute(f"""SELECT * FROM Rates WHERE product_id LIKE "{i[0]}" """)
                myresult = mycursor.fetchall()
                if myresult != []:
                    mycursor.execute(f"""UPDATE Rates SET rate={i[1]} WHERE product_id="{i[0]}" """)
                else:
                    mycursor.execute(
                        f"""INSERT INTO Rates (product_id, rate, scope) VALUES ("{i[0]}", {i[1]}, "{i[2]}")""")
            except:
                print("something went wrong check it")
        return "Ok"


@app.route('/api/trucks', methods=['GET', 'POST', ])
def trucks():
    if request.method == 'POST':
        prov_id = request.form['Provider-Id']
        truck_id = request.form['Truck-Id']
        cursor = billingdb.cursor()
        cursor.execute("USE billdb")
        cursor.execute(f"SELECT id FROM Provider WHERE id='{str(prov_id)}'")
        results = cursor.fetchall()
        if results:
            cursor.execute(f"INSERT INTO Trucks(id, provider_id) VALUES('{str(truck_id)}', '{str(prov_id)}')")
            return Response("Ok", mimetype='text/plain')
        else:
            return Response(f"Provider {prov_id} not found - please enter provider to the providers list",
                            mimetype='text/plain')

    if request.method == 'GET':
        return Response("Please enter truck license plate and provider id:", mimetype='text/plain')


@app.route('/trucks/<truck_id>', methods=['PUT'])
def trucks2(truck_id):
    prov_id = request.form['provider_id']
    mycursor = billingdb.cursor()
    mycursor.execute("USE billdb")
    mycursor.execute(f"SELECT id FROM Provider WHERE id='{str(prov_id)}'")
    results = mycursor.fetchall()
    if results:
        mycursor.execute(f"DELETE FROM Trucks WHERE id='{str(truck_id)}'")
        mycursor.execute(f"INSERT INTO Trucks(id, provider_id) VALUES('{str(truck_id)}', '{str(prov_id):}')")
        return Response(f"changed truck number {truck_id} to provider {prov_id}", mimetype='text/plain')
    else:
        return Response(f"Provider {prov_id} not found -please enter provider to the providers list",
                        mimetype='text/plain')


@app.route('/trucks/<truck_id>/')
def trucktimes(truck_id):
    time1 = request.args.get('from')
    time2 = request.args.get('to')
    # timetest1 = len(time1)
    # timetest2 = len(time2)
    # if timetest1 == 14:
    #     print("good time")
    # else:
    #     timestart = datetime.today().replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # if timetest2 == 14:
    #     print("goodtime")
    # else:
    #     lastday = calendar.monthrange(timestart.year, timestart.month)[1]
    #     timeend = datetime.today().replace(day=lastday, hour=0, minute=0, second=0, microsecond=0)

    payload = {"from": time1, "to": time2}
    res = requests.get(f"http://localhost:8081/testserver/{truck_id}", params=payload)
    if res.status_code == 404:
        return Response({"404"}, status=404)
    return res.json()


# json
@app.route('/bills/<provider_id>/')
def bill(provider_id):
    truck_counter = 0
    session_count = 0
    total = 0
    product_dic = {}
    products = {}
    mycursor = billingdb.cursor()
    mycursor.execute(f"""SELECT name FROM Providr WHERE id='{provider_id}'""")
    name = mycursor.fetchall()[0]
    time1 = request.args.get('from')  # check times
    time2 = request.args.get('to')
    payload = {"from": time1, "to": time2}
    mycursor = billingdb.cursor()
    mycursor.execute("USE billdb")
    mycursor.execute(f"SELECT id FROM Trucks WHERE id='{str(provider_id)}'")
    result = mycursor.fetchall()
    if result:
        for truck in result:
            res = requests.get(f"http://localhost:8081/item/{truck}", params=payload)
            if res.text == "Weight data is unavailable at the  # check what item respons if empty":
                truck_counter += 1
                sessions = res.json()['sessions']
                session_count += len(sessions)
                for session in sessions:
                    r_session = requests.get(f"http://localhost:8081/session/{session}").json()
                    product = r_session['product']
                    neto = int(r_session['neto'])
                    if product in product_dic:
                        product_dic[product]['amount'] += neto
                        product_dic[product]['count'] += 1
                    else:
                        mycursor.execute(
                            f"""SELECT Rate FROM Rates WHERE Scope='{provider_id}'AND Product='{product}' """)
                        rate = int(mycursor.fetchall()[0])
                        if result:
                            product_dic.update({product: {'amount': neto, 'count': 1, 'rate': rate}})
                        else:
                            mycursor.execute(f"""SELECT Rate FROM Rates WHERE Scope='All', Product = '{product}'""")
                            rate = int(mycursor.fetchall()[0])
                            product_dic.update({product: {'amount': neto, 'count': 1, 'rate': rate}})
                for fruit in product_dic:
                    pay = fruit['rate'] * fruit['amount']
                    fruittoadd = {"product": fruit,
                                  "count": fruit['count'],
                                  "amount": fruit['amount'],
                                  "rate": fruit['rate'],
                                  "pay": pay}
                    total += pay
                    products.update(fruittoadd)
                billjson = {
                    "id": provider_id,
                    "name": name,
                    "from": str(time1),
                    "to": str(time2),
                    "truckCount": truck_counter,
                    "sessionCount": session_count,
                    "products": products,
                    "total": total
                }
            else:
                continue
    else:
        return Response(f"Provider {provider_id} not found - please enter provider to the providers list",
                        mimetype='text/plain', status=400)


@app.route("/testserver/123")  # TEST, THIS IS NOT PART OF OUR PROJECT ONLY TEST!!!!!!
def tiesto():
    time1 = request.args.get('from')
    time2 = request.args.get('to')
    if time1 == "10" and time2 == "10":
        return {"id": 123, "tara": 80, "sessions": [1, 4, 6, 8]}
    else:
        return "no good"


if __name__ == '__main__':
    try:
        billingdb = mysql.connector.connect(
            host="billingdb",
            user="root",
            password="1234!",
        )
        mycursor = billingdb.cursor()
        mycursor.execute("use billdb")
        billingdb.commit()

    except:
        print("failed to connect to DB")

    # HOST="billingdb"
    # COMMAND="mysqldump -u root -p 1234! -p billdb > /db/newdb.sql"
    # ssh = subprocess.Popen(["ssh", "%s" % HOST, COMMAND],
    # shell=False,
    # stdout=subprocess.PIPE,
    # stderr=subprocess.PIPE)

    app.run(debug=True, port=8081, host='0.0.0.0')
