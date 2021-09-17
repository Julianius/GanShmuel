from flask import Flask, render_template, request, jsonify
from flask import Response
import mysql.connector
import json
import requests
import random
import os
import openpyxl
import calendar
from datetime import datetime
from werkzeug.wrappers import response
import subprocess
import sys

app = Flask(__name__)


@app.route('/')
def index():
    return render_template('index.html')
    # Add Navigiation bar to our APIs


@app.route('/health')
def health():
    try:
        billingdb = mysql.connector.connect(
            host="billingdb",
            user="root",
            password="1234!",
            database='billdb',
        )
        return Response({"Ok"}, status=200)
    except:
        return Response({"Internal server error"}, status=500)


@app.route('/providers')
def provider():
    return render_template('providers.html')


@app.route('/rates')
def rates():
    mydir = os.listdir("/in")
    return render_template('rates.html', mydir=mydir)


@app.route('/trucks')
def truck():
    return render_template('trucks.html')


@app.route('/trucks/<truck_number>')
def truck_id(truck_number):
    return render_template('truck_id.html')


@app.route('/api/providers', methods=['GET', 'POST'])
def providers():
    # billingdb=connecttosql()
    if request.method == 'POST':
        provider = request.form['provider']
        mycursor = billingdb.cursor()
        mycursor.execute("USE billdb")
        mycursor.execute(f"SELECT name from Provider where name='{str(provider)}'")
        results = mycursor.fetchall()
        if not results:
            switch = True
            while switch:
                prov_id = random.randint(1, 999999)
                jsonprovider = {'id': str(prov_id), 'name': str(provider)}
                try:
                    mycursor.execute("USE billdb")
                    mycursor.execute(f"INSERT INTO Provider(id, name) VALUES('{str(prov_id)}', '{str(provider)}')")
                    switch = False
                except mysql.connector.errors.IntegrityError:
                    continue
                except:
                    return Response('error', status=400)
                return jsonify(jsonprovider)
        else:
            return Response('name exist', status=400)
    if request.method == 'GET':
        return Response("enter provider name:", mimetype='text/plain')


@app.route('/api/rates', methods=['POST', 'GET'])
def ratespost():
    if request.method == 'GET':
        mycursor = billingdb.cursor()
        mycursor.execute("""select * from Rates""")
        row_headers = [x[0] for x in mycursor.description]
        rv = mycursor.fetchall()
        json_data = []
        for result in rv:
            json_data.append(dict(zip(row_headers, result)))
        jsonout = json.dumps(json_data)
        return jsonout  # Check with chris if the JSON type is required on response
    if request.method == 'POST':
        mycursor = billingdb.cursor()
        filename = request.form['msgfile']
        try:
            wrkbk = openpyxl.load_workbook(f"/in/{filename}")
        except openpyxl.utils.exceptions.InvalidFileException:
            return "no such a file"

        sh = wrkbk.active
        semilist = []
        mylist = []
        for i in range(2, sh.max_row + 1):
            mylist.append(semilist)
            semilist = []
            for j in range(1, sh.max_column + 1):
                cell_obj = sh.cell(row=i, column=j)
                semilist.append(str(cell_obj.value))
        mylist.pop(0)
        for i in mylist:
            try:
                mycursor.execute(f"""SELECT * FROM Rates WHERE product_id LIKE "{i[0]}" """)
                myresult = mycursor.fetchall()
                if myresult != []:
                    mycursor.execute(f"""UPDATE Rates SET rate={i[1]} WHERE product_id="{i[0]}" """)
                else:
                    mycursor.execute(
                        f"""INSERT INTO Rates (product_id, rate, scope) VALUES ("{i[0]}", {i[1]}, "{i[2]}")""")
            except:
                print("something went wrong check it")
        return "Ok"


@app.route('/api/trucks', methods=['GET', 'POST', ])
def trucks():
    if request.method == 'POST':
        prov_id = request.form['Provider-Id']
        truck_id = request.form['Truck-Id']
        cursor = billingdb.cursor()
        cursor.execute("USE billdb")
        cursor.execute(f"SELECT id FROM Provider WHERE id='{str(prov_id)}'")
        results = cursor.fetchall()
        if results:
            try:
                cursor.execute(f"INSERT INTO Trucks(id, provider_id) VALUES('{str(truck_id)}', '{str(prov_id)}')")
                return Response("Ok", mimetype='text/plain')
            except mysql.connector.errors.IntegrityError:
                return Response("truck id all ready exist", status=500)

        else:
            return Response(f"Provider {prov_id} not found - please enter provider to the providers list",
                            mimetype='text/plain')

    if request.method == 'GET':
        return Response("Please enter truck license plate and provider id:", mimetype='text/plain')


@app.route('/trucks/<truck_id>', methods=['PUT'])
def trucks2(truck_id):
    prov_id = request.form['provider_id']
    mycursor = billingdb.cursor()
    mycursor.execute("USE billdb")
    mycursor.execute(f"SELECT id FROM Provider WHERE id='{str(prov_id)}'")
    results = mycursor.fetchall()
    if results:
        mycursor.execute(f"DELETE FROM Trucks WHERE id='{str(truck_id)}'")
        mycursor.execute(f"INSERT INTO Trucks(id, provider_id) VALUES('{str(truck_id)}', '{str(prov_id):}')")
        return Response(f"changed truck number {truck_id} to provider {prov_id}", mimetype='text/plain')
    else:
        return Response(f"Provider {prov_id} not found -please enter provider to the providers list",
                        mimetype='text/plain')


@app.route('/truck/<truck_id>')
def trucktime(truck_id):
    time1 = request.args.get('from')
    time2 = request.args.get('to')
    # timetest1 = len(time1)
    # timetest2 = len(time2)
    # if timetest1 == 14:
    #     print("good time")
    # else:
    #     timestart = datetime.today().replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # if timetest2 == 14:
    #     print("goodtime")
    # else:
    #     lastday = calendar.monthrange(timestart.year, timestart.month)[1]
    #     timeend = datetime.today().replace(day=lastday, hour=0, minute=0, second=0, microsecond=0)

    payload = {"from": time1, "to": time2}
    res = requests.get(f"http://172.28.0.5:5000/item/{truck_id}", params=payload)
    if res.text == "No data found":
        return Response({"404"}, status=404)
    return res.json()

#json
@app.route('/bill/<provider_id>')
def totalbill(provider_id):
    truck_counter = 0
    session_count = 0
    total=0
    product_dic = {}
    products=[]
    mycursor.execute("USE billdb")
    mycursor.execute(f"""SELECT name FROM Provider WHERE id={provider_id}""")
    for names in mycursor.fetchall():
        name = names[0]
    time1 = request.args.get('from') #check times
    time2 = request.args.get('to')
    payload = {"from": time1, "to": time2}
    mycursor.execute(f"SELECT id FROM Trucks WHERE provider_id={provider_id}")
    result = mycursor.fetchall()
    if result:
        for truck in result:
            res = requests.get(f"http://172.28.0.5:5000/item/{truck[0]}", params=payload)
            if res.text != "No data found":
                truck_counter += 1
                sessions1 = res.json()
                sessions=sessions1['sessions']
                session_count += len(sessions)
                # sessions = [3] # not part of the prodaction code give as good session with neto weight
                for session in sessions:
                    r_session = requests.get(f"http://172.28.0.5:5000/session/{session}").json()
                    product = r_session['product_name']
                    neto = int(r_session['neto'])
                    if product in product_dic:
                        product_dic[product]['amount'] += neto
                        product_dic[product]['count'] += 1
                    else:
                        mycursor.execute(f"""SELECT rate FROM Rates WHERE scope={provider_id} AND product_id='{product}' """)
                        rate = mycursor.fetchall()
                        if rate:
                            for rate_list in rate:
                                product_dic.update({product: {'amount': neto, 'count': 1 , 'rate': rate_list[0]}})

                        else:

                            mycursor.execute(f"""SELECT rate FROM Rates WHERE scope='All' AND product_id = '{product}' """)
                            rate = mycursor.fetchall()
                            for rate_list in rate:
                                product_dic.update({product: {'amount': neto, 'count': 1, 'rate': rate_list[0]}})
                for fruit in product_dic:
                    pay = int(product_dic[fruit]['rate']) * int(product_dic[fruit]['amount'])
                    fruittoadd={ "product":fruit,
                                  "count": product_dic[fruit]['count'],
                                  "amount":product_dic[fruit]['amount'],
                                  "rate":product_dic[fruit]['rate'],
                                  "pay":pay}
                    total += pay
                    products.append(fruittoadd)
                billjson={
                              "id": provider_id,
                              "name": name,
                              "from": str(time1),
                              "to": str(time2),
                              "truckCount": truck_counter,
                              "sessionCount": session_count,
                              "products": products,
                              "total": total
                            }
                return billjson
            else:
                return "blabla"
    else:
        return Response(f"Provider {provider_id} not found - please enter provider to the providers list",
                        mimetype='text/plain', status=400)


@app.route("/clear/")  # TEST, THIS IS  PART OF OUR PROJECT clear all the test parameters from database!!!!!!
def clear_databases_test():
    provider_id = request.args.get('provider_id')
    truck_id = request.args.get('truck_id')
    truck_id2 = request.args.get('truck_id2')
    mycursor = billingdb.cursor()
    mycursor.execute("USE billdb")
    mycursor.execute(f"DELETE FROM Provider WHERE id= {int(provider_id)}")
    mycursor.execute(f"DELETE FROM Trucks WHERE id='{str(truck_id)}'")
    mycursor.execute(f"DELETE FROM Trucks WHERE id='{str(truck_id2)}'")
    return "ok"

if __name__ == '__main__':
    connectdb = True
    while connectdb:
        try:
            billingdb = mysql.connector.connect(
                host="billingdb",
                user="root",
                password="1234!",
                database='billdb',
            )
            mycursor = billingdb.cursor()
            connectdb = False
        except:
            connectdb = True

    app.run(debug=True, port=8081, host='0.0.0.0')
