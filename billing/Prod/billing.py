from flask import Flask, render_template, request, jsonify
from flask import Response
import mysql.connector
import json
import requests
import random
import os
import openpyxl
import calendar
from datetime import datetime
from werkzeug.wrappers import response
import subprocess
import sys

app = Flask(__name__)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/bills')
def bills():
    return render_template('bills.html')

@app.route('/bills/<provider_id>')
def bills_spes(provider_id):
    return render_template('bills_spec.html')


@app.route('/health')
def health():
    return render_template('health.html')
    
@app.route('/api/health')
def healtho():
    try:
        mycursor.execute("use billdb")
        return Response({"Connection to the database is healthy"}, status=200)
    except:
        return Response({"Connection to the darabase is not healthy"},status=500)

@app.route('/providers')
def provider():
    return render_template('providers.html')


@app.route('/rates')
def rates():
    mydir = os.listdir("/in")
    return render_template('rates.html', mydir=mydir)


@app.route('/trucks')
def truck():
    return render_template('trucks.html')


@app.route('/trucks/<truck_number>')
def truck_id(truck_number):
    return render_template('truck_id.html')


@app.route('/api/providers', methods=['GET', 'POST'])
def providers():
    if request.method == 'POST':
        provider = request.form['provider']
        mycursor = billingdb.cursor()
        mycursor.execute("USE billdb")
        mycursor.execute(f"SELECT name from Provider where name='{str(provider)}'")
        results = mycursor.fetchall()
        if not results:
            switch = True
            while switch:
                prov_id = random.randint(1, 999999)
                jsonprovider = {'id': str(prov_id), 'name': str(provider)}
                try:
                    mycursor.execute("USE billdb")
                    mycursor.execute(f"INSERT INTO Provider(id, name) VALUES('{str(prov_id)}', '{str(provider)}')")
                    switch = False
                except mysql.connector.errors.IntegrityError:
                    continue
                except:
                    return Response('error', status=400)
                return jsonify(jsonprovider)
        else:
            return Response('name exist', status=400)
    if request.method == 'GET':
        return Response("enter provider name:", mimetype='text/plain')


@app.route('/api/rates', methods=['POST', 'GET'])
def ratespost():
    if request.method == 'GET':
        mycursor = billingdb.cursor()
        mycursor.execute("""select * from Rates""")
        row_headers = [x[0] for x in mycursor.description]
        rv = mycursor.fetchall()
        json_data = []
        for result in rv:
            json_data.append(dict(zip(row_headers, result)))
        jsonout = json.dumps(json_data)
        return jsonout  # Check with chris if the JSON type is required on response
    if request.method == 'POST':
        mycursor = billingdb.cursor()
        filename = request.form['msgfile']
        try:
            wrkbk = openpyxl.load_workbook(f"/in/{filename}")
        except openpyxl.utils.exceptions.InvalidFileException:
            return "no such a file"

        sh = wrkbk.active
        semilist = []
        mylist = []
        for i in range(2, sh.max_row + 1):
            mylist.append(semilist)
            semilist = []
            for j in range(1, sh.max_column + 1):
                cell_obj = sh.cell(row=i, column=j)
                semilist.append(str(cell_obj.value))
        mylist.pop(0)
        for i in mylist:
            try:
                mycursor.execute(f"""SELECT * FROM Rates WHERE product_id LIKE "{i[0]}" """)
                myresult = mycursor.fetchall()
                if myresult != []:
                    mycursor.execute(f"""UPDATE Rates SET rate={i[1]} WHERE product_id="{i[0]}" """)
                else:
                    mycursor.execute(
                        f"""INSERT INTO Rates (product_id, rate, scope) VALUES ("{i[0]}", {i[1]}, "{i[2]}")""")
            except:
                print("something went wrong check it")
        return "Ok"


@app.route('/api/trucks', methods=['GET', 'POST', ])
def trucks():
    if request.method == 'POST':
        prov_id = request.form['Provider-Id']
        truck_id = request.form['Truck-Id']
        cursor = billingdb.cursor()
        cursor.execute("USE billdb")
        cursor.execute(f"SELECT id FROM Provider WHERE id='{str(prov_id)}'")
        results = cursor.fetchall()
        if results:
            try:
                cursor.execute(f"INSERT INTO Trucks(id, provider_id) VALUES('{str(truck_id)}', '{str(prov_id)}')")
                return Response("Ok", mimetype='text/plain')
            except mysql.connector.errors.IntegrityError:
                return Response("truck id all ready exist", status=400)
        else:
            return Response(f"Provider {prov_id} not found - please enter provider to the providers list",
                            mimetype='text/plain')

    if request.method == 'GET':
        return Response("Please enter truck license plate and provider id:", mimetype='text/plain')


@app.route('/trucks/<truck_id>', methods=['PUT'])
def trucks2(truck_id):
    prov_id = request.form['provider_id']
    mycursor = billingdb.cursor()
    mycursor.execute("USE billdb")
    mycursor.execute(f"SELECT id FROM Provider WHERE id='{str(prov_id)}'")
    results = mycursor.fetchall()
    if results:
        mycursor.execute(f"DELETE FROM Trucks WHERE id='{str(truck_id)}'")
        mycursor.execute(f"INSERT INTO Trucks(id, provider_id) VALUES('{str(truck_id)}', '{str(prov_id):}')")
        return Response(f"changed truck number {truck_id} to provider {prov_id}", mimetype='text/plain')
    else:
        return Response(f"Provider {prov_id} not found -please enter provider to the providers list",
                        mimetype='text/plain')


@app.route('/trucks/<truck_id>/')
def trucktimes(truck_id):
    time1 = request.args.get('from')
    time2 = request.args.get('to')
    timetest1 = str(time1)
    timetest2 = str(time2)

    if len(timetest1) == 14:
        try:
            datetime(int(timetest1[0:4]), int(timetest1[4:6]), int(timetest1[6:8]), int(timetest1[8:10]), int(timetest1[10:12]), int(timetest1[12:14]))
        except:
            timestart = datetime.today().replace(day=1)
            timestr1 = str(timestart)
            time1 = timestr1.split()[0].replace("-","") + "000000"
    else:
        timestart = datetime.today().replace(day=1)
        timestr1 = str(timestart)
        time1 = timestr1.split()[0].replace("-","") + "000000"

    if len(timetest2) == 14:
        try:
            datetime(int(timetest2[0:4]), int(timetest2[4:6]), int(timetest2[6:8]), int(timetest2[8:10]), int(timetest2[10:12]), int(timetest2[12:14]))
        except:
            timeend = datetime.today()
            time2 = timeend.strftime("%Y%m%d%H%M%S")
    else:
        timeend = datetime.today()
        time2 = timeend.strftime("%Y%m%d%H%M%S")

    payload = {"from": time1, "to": time2}
    res = requests.get(f"http://18.157.175.199:8083/item/{truck_id}", params=payload)
    if res.text == "No data found":
        return Response({"404"}, status=404)
    return res.json()


@app.route('/bills/<provider_id>/')
def totalbill(provider_id):
    time1 = request.args.get('from') 
    time2 = request.args.get('to')
    timetest1 = str(time1)
    timetest2 = str(time2)

    if len(timetest1) == 14:
        print("good time")
        try:
            datetime(int(timetest1[0:4]), int(timetest1[4:6]), int(timetest1[6:8]), int(timetest1[8:10]), int(timetest1[10:12]), int(timetest1[12:14]))
        except:
            timestart = datetime.today().replace(day=1)
            timestr1 = str(timestart)
            time1 = timestr1.split()[0].replace("-","") + "000000"
    else:
        timestart = datetime.today().replace(day=1)
        timestr1 = str(timestart)
        time1 = timestr1.split()[0].replace("-","") + "000000"

    if len(timetest2) == 14:
        print("good time")
        try:
            datetime(int(timetest2[0:4]), int(timetest2[4:6]), int(timetest2[6:8]), int(timetest2[8:10]), int(timetest2[10:12]), int(timetest2[12:14]))
        except:
            timeend = datetime.today()
            time2 = timeend.strftime("%Y%m%d%H%M%S")
    else:
        timeend = datetime.today()
        time2 = timeend.strftime("%Y%m%d%H%M%S")

    payload = {"from": time1, "to": time2}
    truck_counter = 0
    session_count = 0
    total=0
    product_dic = {}
    products=[]
    mycursor.execute("USE billdb")
    mycursor.execute(f"""SELECT name FROM Provider WHERE id={provider_id}""")
    for names in mycursor.fetchall():
        name = names[0]
    mycursor.execute(f"SELECT id FROM Trucks WHERE provider_id={provider_id}")
    result = mycursor.fetchall()
    if result:
        for truck in result:
            res = requests.get(f"http://18.157.175.199:8083/item/{truck[0]}", params=payload)
            if res.text != "No data found":
                truck_counter += 1
                sessions1 = res.json()
                sessions=sessions1['sessions']
                session_count += len(sessions)
                for session in sessions:
                    r_session = requests.get(f"http://18.157.175.199:8083/session/{session}").json()
                    product = r_session['product_name']
                    neto = int(r_session['neto'])
                    if product in product_dic:
                        product_dic[product]['amount'] += neto
                        product_dic[product]['count'] += 1
                    else:
                        mycursor.execute(f"""SELECT rate FROM Rates WHERE scope={provider_id} AND product_id='{product}' """)
                        rate = mycursor.fetchall()
                        if rate:
                            for rate_list in rate:
                                product_dic.update({product: {'amount': neto, 'count': 1 , 'rate': rate_list[0]}})

                        else:

                            mycursor.execute(f"""SELECT rate FROM Rates WHERE scope='All' AND product_id = '{product}' """)
                            rate = mycursor.fetchall()
                            for rate_list in rate:
                                product_dic.update({product: {'amount': neto, 'count': 1, 'rate': rate_list[0]}})
                for fruit in product_dic:
                    pay = int(product_dic[fruit]['rate']) * int(product_dic[fruit]['amount'])
                    fruittoadd={ "product":fruit,
                                  "count": product_dic[fruit]['count'],
                                  "amount":product_dic[fruit]['amount'],
                                  "rate":product_dic[fruit]['rate'],
                                  "pay":pay}
                    total += pay
                    products.append(fruittoadd)
                billjson={
                              "id": provider_id,
                              "name": name,
                              "from": str(time1),
                              "to": str(time2),
                              "truckCount": truck_counter,
                              "sessionCount": session_count,
                              "products": products,
                              "total": total
                            }
                return json.dumps(billjson)
            else:
                return "Not working"
    else:
        return Response(f"Provider {provider_id} not found - please enter provider to the providers list",
                        mimetype='text/plain', status=400)

@app.route("/clear/")  # TEST, THIS IS  PART OF OUR PROJECT clear all the test parameters from database!!!!!!
def clear_databases_test():
    provider_id = request.args.get('provider_id')
    truck_id = request.args.get('truck_id')
    truck_id2 = request.args.get('truck_id2')
    mycursor = billingdb.cursor()
    mycursor.execute("USE billdb")
    mycursor.execute(f"DELETE FROM Provider WHERE id= {int(provider_id)}")
    mycursor.execute(f"DELETE FROM Trucks WHERE id='{str(truck_id)}'")
    mycursor.execute(f"DELETE FROM Trucks WHERE id='{str(truck_id2)}'")
    return "ok"


if __name__ == '__main__':
    connectdb = True
    counter = 60 
    while connectdb:
        counter -= 1
        time.sleep(1)
        if counter == 0:
            break
        try:
            billingdb = mysql.connector.connect(
                host="billingdb",
                user="root",
                password="1234!",
                database='billdb',
            )
            mycursor = billingdb.cursor()
            connectdb = False
        except:
            connectdb = True

    app.run(debug=True, port=8081, host='0.0.0.0')
