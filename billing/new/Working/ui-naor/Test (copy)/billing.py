import os
from flask import Flask, render_template, request, jsonify
from flask import Response
import mysql.connector
import json
import requests
import openpyxl

ifconnect = True
app = Flask(__name__)

@app.route('/rates')
def rates():
    mydir = os.listdir("./")
    return render_template('rates.html',mydir=mydir)

@app.route('/api/rates', methods=['POST','GET'])
def ratespost():
    if request.method == 'GET':
        mycursor = billingdb.cursor()
        mycursor.execute("""select * from Rates""")
        row_headers=[x[0] for x in mycursor.description] 
        rv = mycursor.fetchall()
        json_data=[]
        for result in rv:
            json_data.append(dict(zip(row_headers,result)))
        jsonout = json.dumps(json_data) 
        return jsonout
    if request.method == 'POST':
        mycursor = billingdb.cursor()
        filename = request.form['msgfile']
        wrkbk = openpyxl.load_workbook(f"/in/{filename}")
        sh = wrkbk.active
        semilist = []
        mylist = []
        for i in range(2, sh.max_row+1):
            mylist.append(semilist)
            semilist=[]
            for j in range(1, sh.max_column+1):
                cell_obj = sh.cell(row=i, column=j)
                semilist.append(str(cell_obj.value))

        mylist.pop(0)
        for i in mylist:
            try:
                mycursor.execute(f"""SELECT * FROM Rates WHERE product_id LIKE "{i[0]}" """)
                myresult = mycursor.fetchall()
                if myresult != []:
                    mycursor.execute(f"""UPDATE Rates SET rate={i[1]} WHERE product_id="{i[0]}" """)
                else:
                    mycursor.execute(f"""INSERT INTO Rates (product_id, rate, scope) VALUES ("{i[0]}", {i[1]}, "{i[2]}")""")

            except:
                print("something went wrong check it")
        return ""
        
@app.route('/')
def index():
    return "Welcome to Billing Blue Site"

@app.route('/health')
def health():
    if ifconnect == True:
        return Response({"Ok"}, status=200)
    else:
        return Response({"Internal server error"}, status=500)

# @app.route('/provider')
# def provider():
#     return render_template('providers.html')

# @app.route('/provider', methods=['GET', 'POST'])
# def providers():
#     if request.method == 'POST':
#         provider = request.form['provider']
#         provid = request.form["provid"]
#         jsonprovider = {"id": provid, "name": provider}
#         with open(f"./templates/providerjson.json", "a+") as chat_file:
#             chat_file.write(str(jsonprovider))
#         billingdb = mysql.connector.connect(
#             host="billingdb",
#             user="root",
#             password="1234!",
#             database='billdb',
#         )
#         mycursor = billingdb.cursor()
#         mycursor.execute("USE billdb")
#         # mycursor.execute(f"SELECT 'name' FROM Provider")
#         # result = mycursor.fetchone()
#         # if provider not in result:
#         mycursor.execute(f"INSERT INTO Provider(id, name) VALUES('{str(provid)}', '{str(provider)}')")
#         return requests.post('localhost:8081/provider', json={'id': 1, 'name': ''})

#     elif request.method == 'GET':
#         return Response({"enter"}, mimetype='text/plain')

if __name__ == '__main__':
    ifconnect = False
    billingdb = mysql.connector.connect(
        host="localhost",
        user="root",
        password="123",
        database="billdb"
    )
    # mycursor = billingdb.cursor()
    # mycursor.execute("CREATE DATABASE IF NOT EXISTS billdb")
    # mycursor.execute("USE billdb")
    # mycursor.execute(
    #     "CREATE TABLE IF NOT EXISTS Provider (id int(11) NOT NULL AUTO_INCREMENT,name varchar(255) DEFAULT NULL,PRIMARY KEY (id)) ENGINE=MyISAM  AUTO_INCREMENT=10001")
    # mycursor.execute(
    #     "CREATE TABLE IF NOT EXISTS Rates (product_id varchar(50) NOT NULL,rate int(11) DEFAULT 0,scope varchar(50) DEFAULT NULL,FOREIGN KEY (scope) REFERENCES Provider (id)) ENGINE=MyISAM")
    # mycursor.execute(
    #     "CREATE TABLE IF NOT EXISTS Trucks (id varchar(10) NOT NULL,provider_id int(11) DEFAULT NULL,PRIMARY KEY (id),FOREIGN KEY (provider_id) REFERENCES Provider (id)) ENGINE=MyISAM")

    # try:
    #     myresult = mycursor.fetchall()
    #     billingdb.commit()
    #     ifconnect = True
    # except:
    #     ifconnect = False

    app.run(debug=True, port=8081, host='0.0.0.0')