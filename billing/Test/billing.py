from flask import Flask, render_template, request, jsonify
from flask import Response
import mysql.connector
import json
import requests
import random
import os 
import openpyxl

app = Flask(__name__)


@app.route('/')
def index():
    return "Welcome to Billing Blue Site"


@app.route('/health')
def health():
    if ifconnect == True:
        return Response({"Ok"}, status=200)
    else:
        return Response({"Internal server error"}, status=500)


@app.route('/provider')
def provider():
    return render_template('providers.html')

@app.route('/rates')
def rates():
    mydir = os.listdir("/in")
    return render_template('rates.html',mydir=mydir)

@app.route('/trucks')
def truck():
    return render_template('trucks.html')


@app.route('/api/provider', methods=['GET', 'POST'])
def providers():
    if request.method == 'POST':
        provider = request.form['provider']
        billingdb = mysql.connector.connect(
            host="billingdb",
            user="root",
            password="1234!",
            database='billdb',
        )
        mycursor = billingdb.cursor()
        mycursor.execute("USE billdb")
        mycursor.execute(f"SELECT name from Provider where name='{str(provider)}'")
        results = mycursor.fetchall()
        if not results:
            switch = True
            while switch:
                prov_id = random.randint(1, 999999)
                jsonprovider = {'id': str(prov_id), 'name': str(provider)}
                try:
                    mycursor.execute("USE billdb")
                    mycursor.execute(f"INSERT INTO Provider(id, name) VALUES('{str(prov_id)}', '{str(provider)}')")
                    switch = False
                except mysql.connector.errors.IntegrityError:
                    continue
                    return Response('id exist', status=400)
                except:
                    return Response('error', status=400)
                return jsonify(jsonprovider)
        else:
            return Response('name exist', status=400)
    elif request.method == 'GET':
        return Response("enter provider name:", mimetype='text/plain')


@app.route('/api/rates', methods=['POST','GET'])
def ratespost():
    if request.method == 'GET':
        mycursor = billingdb.cursor()
        mycursor.execute("""select * from Rates""")
        row_headers=[x[0] for x in mycursor.description] 
        rv = mycursor.fetchall()
        json_data=[]
        for result in rv:
            json_data.append(dict(zip(row_headers,result)))
        jsonout = json.dumps(json_data) 
        return jsonout
    if request.method == 'POST':
        mycursor = billingdb.cursor()
        filename = request.form['msgfile']
        wrkbk = openpyxl.load_workbook(f"/in/{filename}")
        sh = wrkbk.active
        semilist = []
        mylist = []
        for i in range(2, sh.max_row+1):
            mylist.append(semilist)
            semilist=[]
            for j in range(1, sh.max_column+1):
                cell_obj = sh.cell(row=i, column=j)
                semilist.append(str(cell_obj.value))
        mylist.pop(0)
        for i in mylist:
            try:
                mycursor.execute(f"""SELECT * FROM Rates WHERE product_id LIKE "{i[0]}" """)
                myresult = mycursor.fetchall()
                if myresult != []:
                    mycursor.execute(f"""UPDATE Rates SET rate={i[1]} WHERE product_id="{i[0]}" """)
                else:
                    mycursor.execute(f"""INSERT INTO Rates (product_id, rate, scope) VALUES ("{i[0]}", {i[1]}, "{i[2]}")""")
            except:
                print("something went wrong check it")
        return "hello"

@app.route('/api/trucks', methods=['GET', 'POST', 'PUT'])
def trucks():
    if request.method == 'POST':
        prov_id = request.form['Provider-Id']
        truck_id = request.form['Truck-Id']
        billingdb = mysql.connector.connect(
            host="billingdb",
            user="root",
            password="1234!",
            database='billdb',
        )
        cursor = billingdb.cursor()
        cursor.execute("USE billdb")
        cursor.execute(f"SELECT id FROM Provider WHERE id='{str(prov_id)}'")
        results = cursor.fetchall()
        if results:
            cursor.execute(f"INSERT INTO Trucks(id, provider_id) VALUES('{str(truck_id)}', '{str(prov_id)}')")
            return Response("Ok", mimetype='text/plain')
        else:
            return Response("Provider not found - please enter provider to the providers list", mimetype='text/plain')

    elif request.method == 'GET':
        billingdb = mysql.connector.connect(
            host="billingdb",
            user="root",
            password="1234!",
            database='billdb',
        )
        cursor = billingdb.cursor()
        cursor.execute("USE billdb")
        cursor.execute(f"SELECT id FROM Provider")
        return Response("Please enter truck license plate and provider id:", mimetype='text/plain')



if __name__ == '__main__':
    try:
        ifconnect = False
        billingdb = mysql.connector.connect(
            host="billingdb",
            user="root",
            password="1234!",
        )
        mycursor = billingdb.cursor()
        mycursor.execute("CREATE DATABASE IF NOT EXISTS billdb")
        mycursor.execute("USE billdb")
        mycursor.execute(
            "CREATE TABLE IF NOT EXISTS Provider (id int(11) NOT NULL AUTO_INCREMENT,name varchar(255) DEFAULT NULL,PRIMARY KEY (id)) ENGINE=MyISAM  AUTO_INCREMENT=10001")
        mycursor.execute(
            "CREATE TABLE IF NOT EXISTS Rates (product_id varchar(50) NOT NULL,rate int(11) DEFAULT 0,scope varchar(50) DEFAULT NULL,FOREIGN KEY (scope) REFERENCES Provider (id)) ENGINE=MyISAM")
        mycursor.execute(
            "CREATE TABLE IF NOT EXISTS Trucks (id varchar(10) NOT NULL,provider_id int(11) DEFAULT NULL,PRIMARY KEY (id),FOREIGN KEY (provider_id) REFERENCES Provider (id)) ENGINE=MyISAM")
        
        myresult = mycursor.fetchall()
        billingdb.commit()
        ifconnect = True
    except:
        ifconnect = False

    app.run(debug=True, port=8081, host='0.0.0.0')
